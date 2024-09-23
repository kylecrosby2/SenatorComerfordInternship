#! python3
# AddSchoolCommitteeMembers.py - Add new school committee member emails to CF.

from selenium import webdriver
import bs4
import time
import openpyxl as excel


def main():
    path = "C:/Users/kerry/Documents/chromedriver.exe"
    driver = webdriver.Chrome(path)
    driver2 = webdriver.Chrome(path)
    login_cf(driver)
    login_cf(driver2)
    skipped_file = open('SchoolCommitteeMembersNotInCF.txt', 'a', encoding='utf-8')
    wb = excel.load_workbook('SchoolCommitteeList.xlsx')
    sheet = wb['School Committees']
    member_dict_list = get_data_from_sheet(sheet)
    cf_page(member_dict_list, driver, driver2, skipped_file)
    skipped_file.close()


def get_data_from_sheet(sheet):
    member_dict_list = []
    for i in range(2, 136):
        member_dict = {'Name': sheet.cell(row=i, column=2).value, 'Email': sheet.cell(row=i, column=4).value, 'Position': sheet.cell(row=i, column=3).value, 'Town': sheet.cell(row=i, column=2).value}
        member_dict_list.append(member_dict)
    return member_dict_list


def cf_page(member_dict_list, driver, driver2, skipped_file):
    for d in member_dict_list:
        try:
            driver.refresh()
            driver.find_element_by_id('call-log-tab').click()
            name_elem_num = get_constituent_number(driver, d, skipped_file)
            driver2.get('https://communityfluency.com/office/constituents/%s' % name_elem_num)
            add_email(driver2, d['Email'])
            add_to_constituent_groups(d, driver2, name_elem_num)
        except Exception as exc:
            print('Error: %s. %s will be skipped.' % (exc, d['Name']))
            skipped_file.write(str(d))
            continue


def add_email(driver2, email):
    driver2.find_element_by_xpath('//*[@id="main"]/div[1]/a/button').click()
    soup = bs4.BeautifulSoup(driver2.page_source, 'html.parser')
    primary_email = soup.select('#contact_form > table > tbody > tr:nth-child(6) > td:nth-child(2) > div:nth-child(1) > input')[0]
    if primary_email.get('value') is not None or primary_email.get('value') != email:
        driver2.find_element_by_xpath('//*[@id="contact_form"]/table/tbody/tr[6]/td[2]/div[2]/input').send_keys(email)
    else:
        driver2.find_element_by_xpath('//*[@id="contact_form"]/table/tbody/tr[6]/td[2]/div[3]/input[1]').send_keys(email)
    driver2.find_element_by_xpath('//*[@id="contact_form"]/div[2]/input[2]').click()


def add_to_constituent_groups(member_dict, driver, name_elem_num):
    driver.get('https://communityfluency.com/office/constituents/%s/category/94/new' % name_elem_num)
    soup = bs4.BeautifulSoup(driver.page_source, 'html.parser')
    checkbox_elem = soup.find(id='3712')
    if 'keep' not in checkbox_elem.get('name'):
        driver.find_element_by_xpath('//*[@id="3712"]').click()
        driver.find_element_by_xpath('//*[@id="pivot_3712"]/div[1]/input').send_keys('%s, %s' % (member_dict['Position'], member_dict['Town']))
    driver.find_element_by_xpath('//*[@id="main"]/form/div[2]/input[2]').click()


# Logs into cf with the webdriver.
def login_cf(driver):
    driver.get("https://communityfluency.com/office")
    email = driver.find_element_by_id('email')
    password = driver.find_element_by_id('password')
    # Must input your email and password into the quotations below.
    email.send_keys('akylecrosby@gmail.com')
    password.send_keys('$omewhere391Flag')
    sign_in = driver.find_elements_by_xpath('//button[@type="submit"]')[0]
    sign_in.click()
    driver.find_element_by_id('call-log-tab').click()


def get_constituent_number(driver, member_dict, skipped_file):
    finish = True
    name_elem_num = ''
    name = member_dict['Name']
    name_field = driver.find_element_by_id('call-subject')
    # name_field.clear()
    for i in list(name):
        name_field.send_keys(i)
    time.sleep(2)
    try:
        element = driver.find_element_by_xpath('//*[@id="call-list"]/table/tbody/tr[1]/td[1]/span')
        driver.execute_script("$(arguments[0]).click();", element)
    except Exception as exc:
        print('Error: %s. %s will be skipped.' % (exc, name))
        skipped_file.write(str(member_dict))
        driver.refresh()
        driver.find_element_by_id('call-log-tab').click()
        finish = False
    if finish is True:
        html_text = driver.page_source
        soup = bs4.BeautifulSoup(html_text, 'html.parser')
        elems = soup.select('.ml-2')
        name_elem = elems[5]
        name_elem_num = ''
        start_value = ''
        append = False
        for i in list(str(name_elem)):
            start_value += i
            if start_value == '<label class="ml-2" for="add-person-':
                append = True
                continue
            if append is True:
                if i == '>':
                    break
                name_elem_num += i
    return name_elem_num[:-1]


main()
