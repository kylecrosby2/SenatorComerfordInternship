#! python3
# OfficialsNotInCF.py - Determines which district officials aren't in CF based on a comparision of Excel sheets.

import openpyxl as excel
import re
import os
import sys


def compare():
    official_list = excel.load_workbook('EmailListsOfficials.xlsx')
    #constituent_list = excel.load_workbook('Amherst_Bernardston_Colrain.xlsx')
    #constituent_list_sheet = constituent_list[constituent_list.sheetnames[0]]
    for s in official_list.sheetnames:
        value_dict = {'Town': '', 'Name': '', 'Position': '', 'Phone Number': '', 'Address': '', 'Email': ''}
        library_name_list = ['Library', 'Library Director', 'Cell']
        sheet = official_list[s]
        row_num = 0
        while True:
            row_num += 1
            if sheet.cell(row=row_num, column=2).value is None:
                break
        for i in range(1, 7):
            cell = sheet.cell(row=1, column=i).value
            for value in library_name_list:
                if cell is not None:
                    if value.lower() in cell.lower():
                        if value == 'Library':
                            value_dict['Town'] = i
                            continue
                        if value == 'Library Director':
                            value_dict['Name'] = i
                            continue
                        if value == 'Cell':
                            value_dict['Phone Number'] = i
                            continue
            for value in value_dict:
                if cell is not None:
                    if value.lower() in cell.lower() and value_dict[value] == '':
                        value_dict[value] = i
        for i in range(1, row_num):
            # This will loop through the constituent list to find a match.
            name1 = sheet.cell(row=i, column=value_dict['Name']).value
            name = sheet.cell(row=i, column=value_dict['Name']).value
            email = sheet.cell(row=i, column=value_dict['Email']).value
            name = str(name)
            remove_non_alpha = re.compile('[^A-Za-z0-9 ]+')
            remove_roman_numerals = re.compile('^([0-9]+)|([IVXLCM]+)\\.?$')
            name = remove_non_alpha.sub('', name)
            name = remove_roman_numerals.sub('', name)
            name_no_middle = '%s %s' % (name.split(' ')[0], name.split(' ')[-1])
            in_cf = False
            files = ['Amherst-Erving', 'Gill-Leverett', 'Leyden-Northampton', 'Northfield-Shutesbury', 'SouthHadley-Whately']
            for filename in files:
                wb = excel.load_workbook("C:/Users/kerry/PycharmProjects/KylePythonProjects/ComerfordPrograms/TownExcelSheets/" + filename + '.xlsx')
                constituent_sheet = wb[wb.sheetnames[0]]
                row_num1 = 0
                while True:
                    row_num1 += 1
                    if constituent_sheet.cell(row=row_num1, column=1).value is None:
                        break
                for row in range(1, row_num1):
                    #print(constituent_sheet.cell(row=row, column=1).value)
                    # Note: Change the 'Town, int' thing. Also change value_dict into a dictionary.
                    try:
                        if email.lower() == constituent_sheet.cell(row=row, column=4).value.lower() or email.lower() == constituent_sheet.cell(row=row, column=5).value.lower():
                            in_cf = True
                            break
                    except AttributeError:
                        pass
                    if in_cf is False:
                        if name.lower().strip() == str(constituent_sheet.cell(row=row, column=1).value).lower().strip() or name_no_middle.lower().strip() == str(constituent_sheet.cell(row=row, column=1).value).lower().strip() or name.lower().strip() == str(constituent_sheet.cell(row=row, column=2).value).lower().strip():
                            #if str(sheet.cell(row=i, column=value_dict['Town']).value).lower() in str(constituent_sheet.cell(row=row, column=3).value).lower():
                            in_cf = True
                            break
                    if in_cf is False:
                        if name1.lower().strip() == str(constituent_sheet.cell(row=row, column=1).value).lower().strip() or name_no_middle.lower().strip() == str(constituent_sheet.cell(row=row, column=1).value).lower().strip() or name1.lower().strip() == str(constituent_sheet.cell(row=row, column=2).value).lower().strip():
                            #if str(sheet.cell(row=i, column=value_dict['Town']).value).lower() in str(constituent_sheet.cell(row=row, column=3).value).lower():
                            in_cf = True
                            break
            if in_cf is False:
                s = ''
                s += name
                for v in value_dict:
                    try:
                        s += ', ' + str(sheet.cell(row=i, column=int(value_dict[v])).value)
                    except Exception as exc:
                        stop = exc
                print(s)
                list_file = open('NotInCF.txt', 'a')
                list_file.write(s + '\n')
                list_file.close()


compare()
