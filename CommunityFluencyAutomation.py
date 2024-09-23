#! python3
# CommunityFluencyAutomation.py - Automates the process of logging contacts to Community Fluency.


from selenium import webdriver
from selenium.webdriver.support.ui import Select
from threading import Thread
import time
import openpyxl as excel
from difflib import SequenceMatcher
import bs4
import re
import sys


class EmailThread:

    def __init__(self, constituent_name, date, leg_groups, issue_groups_c, email_type, email_text, email_phone_list, pronouns):
        self.constituent_name = constituent_name
        self.date = date
        self.leg_groups = leg_groups
        self.issue_groups_c = issue_groups_c
        self.email_type = email_type
        self.email_text = email_text
        self.email_phone_list = email_phone_list
        self.pronouns = pronouns


def main():
    legislation_groups = ['"Person with a disability" (S.2554)', '100% Renewable Energy (H2836/S1985)', 'Alternative Health Care (S.2634, S.665, H.3660)', 'Animal Research bills (S.534, H.758, H.764)', 'Au pair bills (H.4705, H.4707)', 'Ban Facial Recognition (S.1385/H.1538)', 'Ban Flavored Tobacco (H.4089/S.2357)', 'Ban Glyphosate bill (H.792)', 'Bodyworks bill (S.168)', 'Breakfast After the Bell (S.267)', 'Breakfast after the Bell (H.4218)', 'Carbon pricing (H.2810)', 'Cats & Dogs Retail Sale Ban bills (S.175/H.800)', 'Central Service Technicians (S2447/H4267)', 'Change the State Flag and Seal (S.1877/H.2776)', 'Civil service exams (H.2292)', 'Construction Defect Claims Condo Owners (H4605)', 'Dark Sky bill (H.2858/S.1937)', 'Decarceration and COVID-19 (H.4652)', 'Distracted Driving bill', 'Economic Development Bond Bill (S.2842, H.4879)', 'Election Day Registration (S.396, H.685, H.636)', 'Emergency Paid Sick Time (S.2701/H.4700)', 'Emergency housing bills (S.2831/H.4878)', 'End of Life Options', 'Environmental Justice (H4264/S453/S464)', 'Equality & Respect in the Leg (S.1898, H.3572)', 'FY21 Budget: Pollinator habitats', 'Factory Worker Protection - COVID19 (SD.2934/H.4738)', 'Fairness in Debt Collection (H.4694/S.2734)', 'Family caregiver tax credit (S.702/H.2608)', 'Farmer overtime (S.2347)', 'Female Genital Mutilation (H4606)', 'Film/TV incentive bills (H.2419/S.1728)', 'Flame Retardant Protection (S.1230/S.2338)', 'Free Phone Calls for Incarcerated People (S.2846)', 'Genocide Education (S2581)', 'Gun Crime Data (H.2405/S.1388)', 'Hazard Pay for Essential Workers/Public Employees (H.4631, HD.5031)', 'Healthy Soils', 'Healthy Youth Act (S. 2459)', 'Homeless families & youth state ID (S.2555)', 'Host Community Agreement bills', 'IT Bond Bill 2020 (H.4708/H.4733/S.2759)', 'Incarcerated people protection (S.1372, S.2662, S.1391/H.2127, H.4607)', "Indigenous Peoples' bills", 'Juneteenth Holiday (SD2975/HD5141)', 'Life Without Parole (S.826/H.3358)', 'MCAS COVID moratorium (SD.2986)', 'MassHealth estate recovery (SD.2976/HD.5144)', 'Mosquito/arbovirus (H.4650)', 'Net Metering Exemption for Solar Energy Generation (H2866)', 'Net Zero Emissions by 2050 (H3983)', 'Net Zero Stretch Energy Code', 'Nurse Licensure Compact (H.1944/S.103)', 'Nurse Practitioner full practice authority (S.1330, H.1867)', 'Nurse w/out Supervising MD (S.1157)', 'Out of hospital birth access and safety/midwives (S2863)', 'POST (H.2146)', 'Parents Running for Office (S.408)', 'Pesticides & School Children bill (H.791)', 'Plastic Bag Bill', 'Police Reform (S.2800, S.2820)', 'Pollinators (S.447/H.776, H.763)', 'Post Election Audit bills (S.388, H.694, H.687, H.721)', 'Prevent Death and Disability from Stroke (S2835)', 'Prevent Nuclear Weapons Use (S.2165)', 'Prevent Overdose Deaths, Increase Access to Treatment (H.1712/S.2717)', 'Prohibit Native American Mascots (S.247/H.443)', 'Protect Animals from Abusers bill (H.3772/S.2494)', 'Putting patients first (S.2769)', 'ROE Act (S. 1209)', 'Recess (S.330/H.426)', 'Recycling (H.2837/S.1939)', 'Recycling nips (H.2881/S.452)', 'Reduce racial disparities in maternal health (S.2697/H.4448)', 'Right to Repair bills', 'Right to counsel (S.913/H.3456 & H.1537)', 'SAVE Act NPs (H.1867/S.1330)', 'SAVE Students (S.285)', 'Safe Communities Act (S.1401)', 'Safe Drinking Water in Schools (S2503)', 'Safe Injection/Consumption Sites (H4723)', 'Sample bill (H.123)', 'Saving black lives and transforming public safety (HD5128/SD2968)', 'Secure civil rights through the courts of the Commonwealth (H3277)', 'Senate climate bills (S.2476, S.2477, S.2478)', 'Sewage Pollution in Public Waters (H4921/S490)', 'Sexual Assault on Campuses bills', 'Sexual harassment policy bills (S.929)', 'Social Equity Fund in CCC (S2650)', 'Spouses as caregivers (S.28)', 'Structural racism commission (H.1440)', 'Student Opportunity Act', 'TAFDC and EAEDC Emergency Payments (H.4622)', 'Traffic enforcement cameras (S.2553)', 'Transfer Tax (H.1769)', 'Transportation Revenue bills (H.4506, H.4508)', 'Transportation bond bill (S.1813)', 'Traveling Exhibit Animal bills', 'Universities & medicated abortion (H.3841)', 'Vaccine Community Immunity bills (S.2359, H.4096)', 'Vaccine HPV bill (S.1264)', 'Vaccine Religious Exemption bill (H.3999)', 'Vaccine bill H.1848', 'Voting by Mail and Early Voting in 2020 (S.2654/HD.5026, S.2653, S.2608, H.4737)', 'Wendell Forest bill (H.897)', 'Work & Family Mobility Act']
    issue_groups = ['5G Technology', 'Agriculture, Farms & Farmers', 'Animal Rights', 'Arts & Culture', 'Ban Solitary Confinement', 'Broadband connectivity', 'COVID-19: Add business as essential service', 'COVID-19: Adjunct faculty health care', 'COVID-19: Broadband connectivity', 'COVID-19: Cancel Film & TV Production Incentive Sunset', 'COVID-19: Create extra RAFT funding', 'COVID-19: Early retirement (public schools)', 'COVID-19: Emergency paid sick time', 'COVID-19: Eviction/foreclosure moratorium', 'COVID-19: Food Delivery Commission Cap', 'COVID-19: Local case reporting', 'COVID-19: Mandatory Masks', 'COVID-19: Natural treatments', 'COVID-19: No excuse mail-in voting', 'COVID-19: Non-consent to Contact Tracing', 'COVID-19: PUA', 'COVID-19: Public Education Funding', 'COVID-19: Public Health Data for Nursing Homes', 'COVID-19: Remote Start to K-12 Fall 2020 Semester', 'COVID-19: Reopening', 'COVID-19: Reopening Schools (Early Ed & Care)', 'COVID-19: Reopening Schools (Higher ed)', 'COVID-19: Reopening Schools (K-12)', 'COVID-19: Rideshare Fee Increase', 'COVID-19: ServiceNet Union Contract', 'COVID-19: Small business', 'COVID-19: Testing', 'COVID-19: Transportation Investment', 'COVID-19: UI (and PEUC)', 'COVID-19: UMass FY21 Budget Cuts', 'COVID-19: UMass PSU/USA fall negotiations', 'COVID-19: UMass Reopening Public Health', 'COVID-19: UMass employee asks', 'COVID-19: Unemployment expansion (NOT self-employed)', 'COVID-19: Unemployment expansion for self-employed', 'Civil & Human Rights', 'Close Corporate Tax Loopholes', 'Convention Of States', 'Criminal Justice Reform', 'Democracy & Transparency', 'Disability Rights', 'Divest state pension funds from fossil fuels', 'EEE', 'East Northfield Water Company', 'Education (Early Ed & Care)', 'Education (Higher Ed)', 'Education (K-12)', 'Education (Rural)', 'Education (SPED)', 'Elders', 'Elections/Voting', 'Electric School Buses', 'Energy & Environment', 'FY20 Budget', 'FY20 Supplemental Budget: TAFDC and EAEDC Short Term Relief', 'FY21 Budget', 'FY21 Budget: Ch. 257', 'FY21 DDS funding for day/work programs', 'FY23 Budget', 'Fair Share Amendment', 'Farmers OT (overtime)', 'Farren Care Center closure', 'Firearm safety', 'Fix charter school funding', 'Food Security', 'Foster Care', 'General Pierce Bridge', 'Greenfield RMV', 'Gun Safety', 'Harm Reduction', 'Health Care', 'Hemp', "Holyoke Soldiers' Home funding", 'Housing', 'Immigrant rights', 'Indigenous Peoples', 'Lake Wyola', 'Libraries', 'Lower Flags for International Overdose Awareness Day', 'MRF Contract', 'Medicare For All', 'Mental Health', 'Municipal Power Aggregation', 'National Guard activated by Governor', 'National Popular Vote', 'Neonicotinoids Ban', 'Net neutrality', 'Northampton artifacts/MassDOT project', 'Orange Schools Funding Cuts', 'PACE (Program of All-inclusive Care for the Elderly)', 'Pesticide Use', 'Police reform', 'Polystyrene Ban', 'Prop. 2 1/2', 'Public Health', 'Rail', 'Ranked Choice Voting', 'Real time sales tax', 'Regional & Economic Development', 'Reparations', 'Revenue', 'SEVP changes & international students', 'SMART Program (Solar)', 'Save Wendell State Forest', 'Socially Responsible State Investing', 'Solar', 'Souza Baranowski Correctional Center', 'State Mandated Flu Vaccine for Students', 'Statewide Popular Vote for President', 'Synthetic Nitrogen Fertilizers', 'TCI - Transportation Climate Initiative', 'Test 3', 'Transportation', 'Trash and Recycling Reform', 'UMass', 'Vaping & Tobacco', 'Vaping Ban', 'Veterans', "Women's Health"]
    short_month_list = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    short_day_list = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
    thread_list = get_data_from_sheet()
    path = "C:/Users/kerry/Documents/chromedriver.exe"
    driver = webdriver.Chrome(path)
    driver2 = webdriver.Chrome(path)
    driver3 = webdriver.Chrome(path)
    quick_login(driver, driver2, driver3)
    for thread in thread_list:
        try:
            thread_object = EmailThread('', '', [], [], '', '', [], '')
            individual_emails = split_thread(thread, short_day_list, short_month_list)
            get_groups(legislation_groups, issue_groups, thread_object, individual_emails)
            for i in range(1, len(individual_emails)):
                get_constituent_name(individual_emails, thread_object)
                thread_object.email_text = individual_emails[i]
                email = thread_object.email_text
                thread_object.email_phone_list = find_emails_and_phone_numbers(individual_emails, short_day_list, short_month_list)
                thread_object.pronouns = pronoun_search(individual_emails, short_day_list, short_month_list)
                get_date(email, short_day_list, short_month_list, thread_object)
                thread_object.email_type = determine_email_type(email, short_day_list, short_month_list)
                skip = check_info(thread_object, individual_emails)
                if skip is False:
                    name_elem_num = input_to_cf_log(thread_object, email, driver)
                    quick_edit_profile(driver2, driver3, thread_object, name_elem_num)
                    add_thread_to_sheet(thread, 'Logged Threads')
                    change_to_yes(thread)
                else:
                    add_thread_to_sheet(thread, 'Skipped Threads')
                    break
        except Exception as exc:
            add_thread_to_sheet(thread, 'Skipped Threads')
            print('Error: %s. Thread has been added to skipped sheet to log manually.' % exc)


def quick_login(driver, driver2, driver3):
    if __name__ == '__main__':
        Thread(target=login_cf, args=[driver]).start()
        Thread(target=login_cf, args=[driver2]).start()
        Thread(target=login_cf, args=[driver3]).start()


# Logs into cf with the webdriver.
def login_cf(driver):
    driver.get("https://communityfluency.com/office")
    email = driver.find_element_by_id('email')
    password = driver.find_element_by_id('password')
    # Must input your email and password into the quotations below.
    email.send_keys('akylecrosby@gmail.com')
    password.send_keys('$omewhere391Flag')
    sign_in = driver.find_elements_by_xpath('//button[@type="submit"]')[0]
    sign_in.click()
    driver.find_element_by_id('call-log-tab').click()


# Collects all needed data from the SharePoint Excel sheet.
def get_data_from_sheet():
    wb = excel.load_workbook('CF5.xlsx')
    if len(wb.sheetnames) == 1:
        wb.create_sheet('Skipped Threads')
        wb.create_sheet('Logged Threads')
        wb.save('C:/Users/kerry/PycharmProjects/KylePythonProjects/ComerfordPrograms/CF5.xlsx')
    sheet = wb['CF5']
    row_num = 0
    while True:
        row_num += 1
        if sheet.cell(row=row_num, column=3).value is None:
            break
    text_list = []
    for i in range(1, row_num):
        name = sheet.cell(row=i, column=3).value
        text = sheet.cell(row=i, column=4).value
        yes_or_no = sheet.cell(row=i, column=5).value
        # Must input your name here. Allows the program to get all of the emails assigned to you.
        if 'Kyle' in name and 'No' in yes_or_no:
            text_list.append(text)
    # Returns a list of the body text from every email thread. Another function will divide the threads into
    # separate emails.
    return text_list


def add_thread_to_sheet(thread, sheet_name):
    wb = excel.load_workbook('CF5.xlsx')
    sheet = wb[sheet_name]
    row_num = 0
    while True:
        row_num += 1
        if sheet.cell(row=row_num, column=1).value is None:
            break
    sheet.cell(row=row_num, column=1).value = thread
    wb.save('CF5.xlsx')


def change_to_yes(thread):
    wb = excel.load_workbook('CF5.xlsx')
    sheet = wb['CF5']
    row_num = 0
    while True:
        row_num += 1
        if sheet.cell(row=row_num, column=3).value is None:
            break
    for i in range(1, row_num):
        text = sheet.cell(row=i, column=4).value
        if thread in text:
            sheet.cell(row=i, column=5).value = 'Yes'
    wb.save('C:/Users/kerry/PycharmProjects/KylePythonProjects/ComerfordPrograms/CF5.xlsx')


# Split the thread of emails into individual emails. Also extracts the text before the email.
def split_thread(thread, short_day_list, short_month_list):
    individual_emails = []
    emails_by_line = thread.split('\n')
    current_email = ''
    index = 0
    for line in emails_by_line:
        index += 1
        if line[:5] == 'From:' or line[3:6] in short_day_list or index == len(emails_by_line) or line[5:8] in short_month_list:
            individual_emails.append(current_email)
            current_email = ''
        current_email += line + '\n'
    return individual_emails


def get_groups(legislation_groups, issue_groups, thread_object, individual_emails):
    above_data = individual_emails[0]
    positions_list = ['Supports', 'Concerned', 'Undecided', 'Opposed']
    position = '(Position)'
    for line in above_data.split('\n'):
        for group in issue_groups:
            if SequenceMatcher(None, line, group).ratio() >= 0.5:
                for p in positions_list:
                    if p.lower() in line.lower():
                        position = p
                thread_object.issue_groups_c.append((group, position))
        for group in legislation_groups:
            if SequenceMatcher(None, line, group).ratio() >= 0.5:
                for p in positions_list:
                    if p.lower() in line.lower():
                        position = p
                thread_object.leg_groups.append((group, position))


def get_constituent_name(individual_emails, thread_object):
    email = individual_emails[1]
    name = ''
    for line in email.split('\n'):
        if line[:5] == 'From:':
            for i in list(line[6:]):
                if i == '<':
                    break
                name += i
            break
    if 'Comerford' in name:
        name = ''
        for line in email.split('\n'):
            if line[:3] == 'To:':
                for i in list(line[4:]):
                    if i == '<':
                        break
                    name += i
                break
    thread_object.constituent_name = name.strip()


def determine_email_type(email, short_day_list, short_month_list):
    name = ''
    email_type = 'Email'
    received = False
    email_line = ''
    for line in email.split('\n'):
        if line[:5] == 'From:':
            email_line = line
            for i in list(line[6:]):
                if i == '<':
                    received = True
                    break
                name += i
            break
        if line[3:6] in short_day_list or line[5:8] in short_month_list:
            email_type = "Email Sent To Constituent"
            break
        else:
            email_type = "Email Received From Constituent"
    if 'Comerford' in name or 'masenate' in email_line:
        email_type = "Email Sent To Constituent"
    else:
        if received is True:
            email_type = 'Email Received From Constituent'
    return email_type


def get_date(email, short_day_list, short_month_list, thread_object):
    month_list = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']
    year_list = ['2018', '2019', '2020', '2021', '2022']
    month = ''
    day = ''
    year = ''
    date_words = ''
    for line in email.split('\n'):
        if line[:5] == 'Date:' or line[5:8] in short_month_list:
            space_count = 0
            for l in list(line):
                if l == ' ':
                    space_count += 1
                if space_count == 5:
                    break
                if space_count >= 2:
                    date_words += l
    for m in month_list:
        if m in date_words or short_month_list[month_list.index(m)] in date_words:
            date_words = date_words.replace(m, '').strip()
            date_words = date_words.replace(short_month_list[month_list.index(m)], '').strip()
            month = str(month_list.index(m) + 1)
            if len(month) == 1:
                month = '0' + month
            break
    for y in year_list:
        if y in date_words:
            year = y
    date_words = date_words.replace(year, '')
    date_words = date_words.replace(',', '').strip()
    if len(date_words) == 1:
        date_words = '0' + date_words
    day = date_words
    thread_object.date = '%s/%s/%s' % (month, day, '2020')


def input_to_cf_log(thread_object, email, driver):
    name = thread_object.constituent_name
    notes = driver.find_element_by_id('call-log-notes')
    notes.clear()
    notes.send_keys(email)
    name_field = driver.find_element_by_id('call-subject')
    date_field = driver.find_elements_by_css_selector('#call-log-add > div.w-full.-mt-2 > div.text-grey.flex.mb-2.w-full > div:nth-child(3) > input')[0]
    date_field.clear()
    date_field.send_keys(thread_object.date)
    dropdown = Select(driver.find_element_by_xpath('//*[@id="call-log-add"]/div[1]/div[1]/div[1]/select'))
    if thread_object.email_type == 'Email Received From Constituent':
        dropdown.select_by_visible_text('Email Received From Constituent')
    if thread_object.email_type == 'Email Sent To Constituent':
        dropdown.select_by_visible_text('Email Sent To Constituent')
    save_button = driver.find_elements_by_xpath('//*[@id="call-log-add"]/div[3]/div[2]/input')[0]
    name_field.clear()
    for i in list(name):
        name_field.send_keys(i)
    time.sleep(5)
    name_elem_num = get_constituent_number(driver)
    save_button.click()
    time.sleep(2)
    return name_elem_num


def get_constituent_number(driver):
    html_text = driver.page_source
    soup = bs4.BeautifulSoup(html_text, 'html.parser')
    elems = soup.select('.ml-2')
    name_elem = elems[5]
    name_elem_num = ''
    start_value = ''
    append = False
    print(name_elem)
    for i in list(str(name_elem)):
        start_value += i
        if start_value == '<label class="ml-2" for="add-person-':
            append = True
            continue
        if append is True:
            if i == '>':
                break
            name_elem_num += i
    return name_elem_num[:-1]


def edit_constituent_file(name_elem_num, thread_object, driver2):
    email_phone_list = thread_object.email_phone_list
    pronouns = thread_object.pronouns
    driver2.get('https://communityfluency.com/office/constituents/' + name_elem_num)
    add_list(driver2)
    if email_phone_list != [[], []] or pronouns is not None:
        driver2.find_element_by_xpath('//*[@id="main"]/div[1]/a/button').click()
        add_email_phone_pronouns(driver2, email_phone_list, pronouns)
        driver2.find_element_by_xpath('//*[@id="contact_form"]/div[2]/input[2]').click()
    time.sleep(1)
    return driver2


def add_list(driver2):
    soup = bs4.BeautifulSoup(driver2.page_source, 'html.parser')
    # #main > div.flex.flex-wrap > div:nth-child(2) > div:nth-child(2) > table > tbody > tr:nth-child(1) > td.p-2 > i
    elem = soup.select('#main > div.flex.flex-wrap > div:nth-child(2) > div:nth-child(2) > table > tbody > tr:nth-child(1) > td.p-2')
    if 'Add to List' in elem[0].text:
        driver2.find_element_by_css_selector('#main > div.flex.flex-wrap > div:nth-child(2) > div:nth-child(2) > table > tbody > tr:nth-child(1) > td.p-2 > a').click()


def add_email_phone_pronouns(driver2, email_phone_list, pronouns):
    soup = bs4.BeautifulSoup(driver2.page_source, 'html.parser')
    primary_email = soup.select('#contact_form > table > tbody > tr:nth-child(6) > td:nth-child(2) > div:nth-child(1) > input')[0]
    primary_phone = soup.select('#contact_form > table > tbody > tr:nth-child(7) > td:nth-child(2) > div:nth-child(1) > input')[0]
    email_list = email_phone_list[0]
    phone_list = email_phone_list[1]
    try:
        if email_list[0]:
            if primary_email.get('value') == "":
                email_input = driver2.find_element_by_xpath('//*[@id="contact_form"]/table/tbody/tr[6]/td[2]/div[1]/input')
                email_input.send_keys(email_list[0])
            else:
                if primary_email.get('value') != email_list[0]:
                    email_input = driver2.find_element_by_xpath('//*[@id="contact_form"]/table/tbody/tr[6]/td[2]/div[3]/input[1]')
                    email_input.send_keys(email_list[0])
    except IndexError:
        pass
    try:
        if phone_list[0]:
            if primary_phone.get('value') == "":
                phone_input = driver2.find_element_by_xpath('//*[@id="contact_form"]/table/tbody/tr[7]/td[2]/div[1]/input')
                phone_input.send_keys(phone_list[0])
            else:
                phone_input = driver2.find_element_by_xpath('//*[@id="contact_form"]/table/tbody/tr[7]/td[2]/div[2]/input[1]')
                phone_input.send_keys(phone_list[0])
    except IndexError:
        pass
    if pronouns is not None:
        notes_input = driver2.find_element_by_xpath('//*[@id="contact_form"]/table/tbody/tr[8]/td[2]/textarea')
        notes_input.send_keys('\n\n\nPronouns: %s' % pronouns)


def pronoun_search(individual_emails, short_day_list, short_month_list):
    pronoun_list = ['he/him/his', 'she/her/hers', 'they/them/theirs']
    for email in individual_emails:
        email_type = determine_email_type(email, short_day_list, short_month_list)
        if email_type == "Email Received From Constituent":
            for pronoun in pronoun_list:
                if pronoun in email.lower():
                    return pronoun


# Function that uses regular expressions to find emails and phone numbers in the constituent emails to add to
# their files.
def find_emails_and_phone_numbers(individual_emails, short_day_list, short_month_list):
    email_regex = re.compile(r'''(
                [<|' ]+
                [a-zA-Z0-9._%+-]+
                @
                [a-zA-Z0-9.-]+
                (\.[a-zA-Z]{2,4})
                [>|' ]
                )''', re.VERBOSE)
    phone_regex = re.compile(r'''(
        (\d{3}|\(\d{3}\))?
        (\s|-|\.)?
        (\d{3})
        (\s|-|\.)
        (\d{4})
        (\s*(ext|x|ext.)\s*(\d{2,5}))?
        )''', re.VERBOSE)
    email_list = []
    phone_num_list = []
    for email in individual_emails:
        email_type = determine_email_type(email, short_day_list, short_month_list)
        if email_type == 'Email Received From Constituent':
            for i in email_regex.findall(email):
                if 'Jo.Comerford@masenate.gov' not in i[0]:
                    e = i[0][1:-1]
                    email_list.append(e)
            for groups in phone_regex.findall(email):
                phone_num = "-".join([groups[1], groups[3], groups[5]])
                if groups[8] != "":
                    phone_num += " x" + groups[8]
                phone_num_list.append(phone_num)
    email_list2 = []
    for i in email_list:
        if i == 'info@email.actionnetwork.org' or 'masenate' in i:
            break
        if list(i)[0] == '<':
            email_list2.append(i[1:])
        else:
            email_list2.append(i)
    email_phone_list = [email_list2, phone_num_list]
    return email_phone_list


def add_to_groups(driver2, thread_object, name_elem_num):
    issue_groups_id_dict = {'2020 Presidential Election': '5199', '5G Technology': '4779', 'Agriculture, Farms & Farmers': '3637', 'Animal Rights': '4149', 'Arts & Culture': '4191', 'Ban Solitary Confinement': '5010', 'Broadband connectivity': '4533', 'COVID-19: Add business as essential service': '4535', 'COVID-19: Adjunct faculty health care': '4728', 'COVID-19: Broadband connectivity': '4532', 'COVID-19: Cancel Film & TV Production Incentive Sunset': '4899', 'COVID-19: Create extra RAFT funding': '4726', 'COVID-19: Early retirement (public schools)': '4776', 'COVID-19: Emergency paid sick time': '4596', 'COVID-19: Eviction/foreclosure moratorium': '4525', 'COVID-19: Food Delivery Commission Cap': '4996', 'COVID-19: Local case reporting': '4537', 'COVID-19: Mandatory Masks': '4939', 'COVID-19: Natural treatments': '4740', 'COVID-19: No excuse mail-in voting': '4539', 'COVID-19: Non-consent to Contact Tracing': '4683', 'COVID-19: PUA': '4589', 'COVID-19: Public Education Funding': '4955', 'COVID-19: Public Health Data for Nursing Homes': '4730', 'COVID-19: Remote Start to K-12 Fall 2020 Semester': '5032', 'COVID-19: Reopening': '4647', 'COVID-19: Reopening Schools (Early Ed & Care)': '4896', 'COVID-19: Reopening Schools (Higher ed)': '4906', 'COVID-19: Reopening Schools (K-12)': '4585', 'COVID-19: Rideshare Fee Increase': '4936', 'COVID-19: ServiceNet Union Contract': '4778', 'COVID-19: Small business': '4531', 'COVID-19: Testing': '4541', 'COVID-19: Transportation Investment': '4925', 'COVID-19: UI (and PEUC)': '4538', 'COVID-19: UMass FY21 Budget Cuts': '4947', 'COVID-19: UMass PSU/USA fall negotiations': '5083', 'COVID-19: UMass Reopening Public Health': '5000', 'COVID-19: UMass employee asks': '4686', 'COVID-19: Unemployment expansion (NOT self-employed)': '4530', 'COVID-19: Unemployment expansion for self-employed': '4542', 'Civil & Human Rights': '4388', 'Close Corporate Tax Loopholes': '3676', 'Convention Of States': '4432', 'Criminal Justice Reform': '4095', 'Democracy & Transparency': '4389', 'Disability Rights': '4387', 'Divest state pension funds from fossil fuels': '4200', 'EEE': '4891', 'East Northfield Water Company': '3634', 'Education (Early Ed & Care)': '3662', 'Education (Higher Ed)': '3668', 'Education (K-12)': '3658', 'Education (Rural)': '4685', 'Education (SPED)': '3630', 'Elders': '4350', 'Elections/Voting': '4383', 'Electric School Buses': '3649', 'Energy & Environment': '3661', 'FY20 Budget': '3624', 'FY20 Supplemental Budget: TAFDC and EAEDC Short Term Relief': '4758', 'FY21 Budget': '4192', 'FY21 Budget: Ch. 257': '4406', 'FY21 DDS funding for day/work programs': '4838', 'FY23 Budget': '4311', 'Fair Share Amendment': '4353', 'Farmers OT (overtime)': '4172', 'Farren Care Center closure': '5094', 'Firearm safety': '4312', 'Fix charter school funding': '4352', 'Food Security': '4390', 'Foster Care': '4793', 'General Pierce Bridge': '4351', 'Greenfield RMV': '5035', 'Gun Safety': '3677', 'Harm Reduction': '3663', 'Health Care': '3635', 'Hemp': '3641', "Holyoke Soldiers' Home funding": '4818', 'Housing': '3622', 'Immigrant rights': '4349', 'Indigenous Peoples': '4648', 'Lake Wyola': '4926', 'Libraries': '4423', 'Lower Flags for International Overdose Awareness Day': '4808', 'MRF Contract': '3981', 'Medicare For All': '3619', 'Mental Health': '3638', 'Municipal Power Aggregation': '4131', 'National Guard activated by Governor': '5055', 'National Popular Vote': '4337', 'Neonicotinoids Ban': '5020', 'Net neutrality': '4405', 'Northampton artifacts/MassDOT project': '4833', 'Orange Schools Funding Cuts': '4792', 'PACE (Program of All-inclusive Care for the Elderly)': '4134', 'Pesticide Use': '4306', 'Police reform': '4760', 'Polystyrene Ban': '5009', 'Prop. 2 1/2': '4426', 'Public Health': '3623', 'Rail': '4228', 'Ranked Choice Voting': '3670', 'Real time sales tax': '4346', 'Regional & Economic Development': '4137', 'Reparations': '5095', 'Revenue': '3636', 'SEVP changes & international students': '4885', 'SMART Program (Solar)': '3655', 'Save Wendell State Forest': '3627', 'Socially Responsible State Investing': '4982', 'Solar': '4214', 'Souza Baranowski Correctional Center': '4354', 'State Mandated Flu Vaccine for Students': '5038', 'Statewide Popular Vote for President': '5112', 'Synthetic Nitrogen Fertilizers': '4404', 'TCI - Transportation Climate Initiative': '3650', 'Test 3': '4663', 'Transportation': '3665', 'Trash and Recycling Reform': '3675', 'UMass': '4946', 'Vaping & Tobacco': '4358', 'Vaping Ban': '3671', 'Veterans': '4238', "Women's Health": '4305'}
    leg_groups_id_dict = {'"Person with a disability"  (S.2554)': '4385', '100% Renewable Energy (H2836/S1985)': '4937', 'Alternative Health Care (S.2634, S.665, H.3660)': '3654', 'Animal Research bills (S.534, H.758, H.764)': '3652', 'Au pair bills (H.4705, H.4707)': '4171', 'Ban Facial Recognition (S.1385/H.1538)': '3659', 'Ban Flavored Tobacco (H.4089/S.2357)': '3674', 'Ban Glyphosate bill (H.792)': '3632', 'Bodyworks bill (S.168)': '3648', 'Breakfast After the Bell (S.267)': '4168', 'Breakfast after the Bell (H.4218)': '4381', 'Carbon pricing (H.2810)': '4199', 'Cats & Dogs Retail Sale Ban bills (S.175/H.800)': '3955', 'Central Service Technicians (S2447/H4267)': '4966', 'Change the State Flag and Seal (S.1877/H.2776)': '4805', 'Civil service exams (H.2292)': '4828', 'Construction Defect Claims Condo Owners (H4605)': '4968', 'Dark Sky bill (H.2858/S.1937)': '4203', 'Decarceration and COVID-19 (H.4652)': '4673', 'Distracted Driving bill': '3625', 'Economic Development Bond Bill (S.2842, H.4879)': '4962', 'Election Day Registration (S.396, H.685, H.636)': '3673', 'Emergency Paid Sick Time (S.2701/H.4700)': '4631', 'Emergency housing bills (S.2831/H.4878)': '4890', 'End of Life Options': '3631', 'Environmental Justice (H4264/S453/S464)': '4938', 'Equality & Respect in the Leg (S.1898, H.3572)': '3656', 'FY21 Budget: Pollinator habitats': '4359', 'Factory Worker Protection - COVID19 (SD.2934/H.4738)': '4724', 'Fairness in Debt Collection (H.4694/S.2734)': '4806', 'Family caregiver tax credit (S.702/H.2608)': '4128', 'Farmer overtime (S.2347)': '4204', 'Female Genital Mutilation (H4606)': '4967', 'Film/TV incentive bills (H.2419/S.1728)': '4218', 'Flame Retardant Protection (S.1230/S.2338)': '3642', 'Free Phone Calls for Incarcerated People (S.2846)': '5011', 'Genocide Education (S2581)': '4997', 'Gun Crime Data (H.2405/S.1388)': '4729', 'Hazard Pay for Essential Workers/Public Employees (H.4631, HD.5031)': '4636', 'Healthy Soils': '4345', 'Healthy Youth Act (S. 2459)': '4159', 'Homeless families & youth state ID (S.2555)': '4386', 'Host Community Agreement bills': '4360', 'IT Bond Bill 2020 (H.4708/H.4733/S.2759)': '4744', 'Incarcerated people protection (S.1372, S.2662, S.1391/H.2127, H.4607)': '4853', "Indigenous Peoples' bills": '3669', 'Juneteenth Holiday (SD2975/HD5141)': '4860', 'Life Without Parole (S.826/H.3358)': '3617', 'MCAS COVID moratorium (SD.2986)': '4524', 'MassHealth estate recovery (SD.2976/HD.5144)': '4893', 'Mosquito/arbovirus (H.4650)': '4645', 'Net Metering Exemption for Solar Energy Generation (H2866)': '4940', 'Net Zero Emissions by 2050 (H3983)': '4941', 'Net Zero Stretch Energy Code': '3628', 'Nurse Licensure Compact (H.1944/S.103)': '3645', 'Nurse Practitioner full practice authority (S.1330, H.1867)': '4380', 'Nurse w/out Supervising MD (S.1157)': '3633', 'Out of hospital birth access and safety/midwives (S2863)': '5026', 'PFAS in Food Packaging (H3839/S1315)': '5122', 'POST (H.2146)': '4827', 'Parents Running for Office (S.408)': '3672', 'Pesticides & School Children bill (H.791)': '3657', 'Plastic Bag Bill': '3629', 'Police Reform (S.2800, S.2820)': '4892', 'Pollinators (S.447/H.776, H.763)': '3618', 'Post Election Audit bills (S.388, H.694, H.687, H.721)': '3664', 'Prevent Death and Disability from Stroke (S2835)': '4998', 'Prevent Nuclear Weapons Use (S.2165)': '4492', 'Prevent Overdose Deaths, Increase Access to Treatment (H.1712/S.2717)': '4809', 'Prohibit Native American Mascots (S.247/H.443)': '4807', 'Protect Animals from Abusers bill (H.3772/S.2494)': '3653', 'Putting patients first (S.2769)': '4829', 'ROE Act (S. 1209)': '3621', 'Recess (S.330/H.426)': '4334', 'Recycling (H.2837/S.1939)': '4201', 'Recycling nips (H.2881/S.452)': '4202', 'Reduce racial disparities in maternal health (S.2697/H.4448)': '4304', 'Right to Repair bills': '3626', 'Right to counsel (S.913/H.3456 & H.1537)': '4604', 'SAVE Act NPs (H.1867/S.1330)': '3620', 'SAVE Students (S.285)': '4361', 'Safe Communities Act (S.1401)': '3616', 'Safe Drinking Water in Schools (S2503)': '4951', 'Safe Injection/Consumption Sites (H4723)': '4900', 'Sample bill (H.123)': '4465', 'Saving black lives and transforming public safety (HD5128/SD2968)': '4787', 'Secure civil rights through the courts of the Commonwealth (H3277)': '4788', 'Senate climate bills (S.2476, S.2477, S.2478)': '4132', 'Sewage Pollution in Public Waters (H4921/S490)': '5004', 'Sexual Assault on Campuses bills': '3660', 'Sexual harassment policy bills (S.929)': '4355', 'Social Equity Fund in CCC (S2650)': '4969', 'Spouses as caregivers (S.28)': '4845', 'Structural racism commission (H.1440)': '4826', 'Student Opportunity Act': '4956', 'TAFDC and EAEDC Emergency Payments (H.4622)': '4757', 'Traffic enforcement cameras (S.2553)': '4384', 'Transfer Tax (H.1769)': '4170', 'Transportation Revenue bills (H.4506, H.4508)': '4413', 'Transportation bond bill (S.1813)': '4904', 'Traveling Exhibit Animal bills': '3651', 'Universities & medicated abortion (H.3841)': '4348', 'Vaccine Community Immunity bills (S.2359, H.4096)': '3950', 'Vaccine HPV bill (S.1264)': '3667', 'Vaccine Religious Exemption bill (H.3999)': '3666', 'Vaccine bill H.1848': '3646', 'Voting by Mail and Early Voting in 2020 (S.2654/HD.5026, S.2653, S.2608, H.4737)': '4632', 'Wendell Forest bill (H.897)': '3643', 'Work & Family Mobility Act': '3639'}
    leg_groups = thread_object.leg_groups
    issue_groups = thread_object.issue_groups_c
    if issue_groups != [] or leg_groups != []:
        driver2.get('https://communityfluency.com/office/constituents/' + name_elem_num + '/category/95/new')
        for issue, position in issue_groups:
            soup = bs4.BeautifulSoup(driver2.page_source, 'html.parser')
            issue_elem = soup.find(id=issue_groups_id_dict[issue])
            issue_elem_name = issue_elem.get('name')
            if 'keep' not in issue_elem_name:
                driver2.find_element_by_id(issue_groups_id_dict[issue]).click()
                dropdown = Select(driver2.find_element_by_name('position_' + issue_groups_id_dict[issue]))
                dropdown.select_by_visible_text(position)
        driver2.find_elements_by_xpath('//*[@id="main"]/form/div[2]/input[1]')[0].click()
        driver2.get('https://communityfluency.com/office/constituents/' + name_elem_num + '/category/96/new')
        for leg_issue, position in leg_groups:
            soup = bs4.BeautifulSoup(driver2.page_source, 'html.parser')
            issue_elem = soup.find(id=leg_groups_id_dict[leg_issue])
            issue_elem_name = issue_elem.get('name')
            if 'keep' not in issue_elem_name:
                driver2.find_element_by_id(leg_groups_id_dict[leg_issue]).click()
                dropdown = Select(driver2.find_element_by_name('position_' + leg_groups_id_dict[leg_issue]))
                dropdown.select_by_visible_text(position)
        driver2.find_elements_by_xpath('//*[@id="main"]/form/div[2]/input[1]')[0].click()


def quick_edit_profile(driver2, driver3, thread_object, name_elem_num):
    if __name__ == '__main__':
        Thread(target=add_to_groups, args=[driver3, thread_object, name_elem_num]).start()
        Thread(target=edit_constituent_file, args=[name_elem_num, thread_object, driver2]).start()


def check_info(thread_object, individual_emails):
    skip = False
    positions_list = ['Supports', 'Concerned', 'Undecided', 'Opposed']
    position = '(Position)'
    while True:
        checklist = [(thread_object.email_text, 'Message'), (thread_object.constituent_name, 'Name'),
                     (thread_object.date, 'Date'), (thread_object.email_type, 'Email Type'),
                     (thread_object.issue_groups_c, 'Issue Groups'), (thread_object.leg_groups, 'Legislation Groups'),
                     (thread_object.email_phone_list[0], 'Emails of constituent'),
                     (thread_object.email_phone_list[1], 'Phone numbers of constituent'),
                     (thread_object.pronouns, 'Pronouns of constituent')]

        index = 0
        print('\n\n')
        print('Notes on top: %s' % individual_emails[0])
        for value, title in checklist:
            index += 1
            print('%s. %s: %s' % (str(index), title, value))
        user_input = input('Press Enter to confirm. If editing is needed, press the corresponding number and then Enter. Press S and then Enter to Skip this thread.')
        if str(user_input).isdigit() is True:
            user_input = str(user_input)
            if 1 <= int(user_input) <= 8:
                print('Editing the %s' % checklist[int(user_input) - 1][1])
                print('Here is the current content: %s' % checklist[int(user_input) - 1][0])
                user_changes = input('Type the new value here, and press Enter: ')
                if user_input == '1':
                    thread_object.email_text = user_changes
                if user_input == '2':
                    thread_object.constituent_name = user_changes
                if user_input == '3':
                    thread_object.date = user_changes
                if user_input == '4':
                    thread_object.email_type = user_changes
                if user_input == '5':
                    user_input1 = str(input('Press 1 to clear the group list or just press Enter to add to the list. '))
                    if user_input1 == '1':
                        thread_object.issue_groups_c.clear()
                    for p in positions_list:
                        if p in user_changes:
                            position = p
                            user_changes = user_changes.replace(p, '').strip()
                    thread_object.issue_groups_c.append((user_changes, position))
                if user_input == '6':
                    user_input1 = str(input('Press 1 to clear the group list or just press Enter to add to the list. '))
                    if user_input1 == '1':
                        thread_object.leg_groups.clear()
                    for p in positions_list:
                        if p in user_changes:
                            position = p
                            user_changes = user_changes.replace(p, '')
                    thread_object.leg_groups.append((user_changes, position))
                if user_input == '7':
                    thread_object.email_phone_list[0].clear()
                    thread_object.email_phone_list[0].append(user_changes)
                if user_input == '8':
                    thread_object.email_phone_list[1].clear()
                    thread_object.email_phone_list[1].append(user_changes)
                if user_input == '9':
                    thread_object.pronouns = user_changes
        elif user_input.lower() != 's':
            break
        if user_input.lower() == 's':
            skip = True
            print('Thread has been skipped. It has been added to a sheet in the Excel Book to be added manually later.')
            break
    return skip


main()
