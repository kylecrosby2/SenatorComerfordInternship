#! python3
# AddEmails2.py - Adds emails of all district officials to CF.

from selenium import webdriver
from selenium.webdriver.support.ui import Select
import openpyxl as excel
import time
import bs4


def main():
    skipped_list = open('SkippedOfficialList.txt', 'a', encoding='utf-8')
    dict_list = get_data_from_excel()
    path = "C:/Users/kerry/Documents/chromedriver.exe"
    driver = webdriver.Chrome(path)
    driver2 = webdriver.Chrome(path)
    login_cf(driver)
    login_cf(driver2)
    add_email_phone(driver, dict_list, driver2, skipped_list)
    skipped_list.close()
    print('Done')


# Logs into cf with the webdriver.
def login_cf(driver):
    driver.get("https://communityfluency.com/office")
    email = driver.find_element_by_id('email')
    password = driver.find_element_by_id('password')
    # Must input your email and password into the quotations below.
    email.send_keys('akylecrosby@gmail.com')
    password.send_keys('$omewhere391Flag')
    sign_in = driver.find_elements_by_xpath('//button[@type="submit"]')[0]
    sign_in.click()
    driver.find_element_by_id('call-log-tab').click()


def get_data_from_excel():
    official_list = excel.load_workbook('EmailListsOfficials1.xlsx')
    dict_list = []
    for s in official_list.sheetnames:
        sheet = official_list[s]
        position_dict = {'Name': 0, 'Email': 0, 'Phone': 0, 'Town': 0, 'Position': 0, 'Title': 0}
        for i in range(1, 7):
            cell = sheet.cell(row=1, column=i).value
            if cell is not None:
                if 'Name' in cell or 'Director' in cell:
                    position_dict['Name'] = i
                if 'Email' in cell:
                    position_dict['Email'] = i
                if 'Phone' in cell or 'Ph #' in cell or 'Notes' in cell:
                    position_dict['Phone'] = i
                if 'Town' in cell or 'Library' in cell:
                    position_dict['Town'] = i
                if 'Position' in cell or 'School' in cell or 'Library' in cell:
                    position_dict['Position'] = i
        row_num = 0
        while True:
            row_num += 1
            if sheet.cell(row=row_num, column=2).value is None:
                break
        for i in range(2, row_num-1):
            info_dict = {'Name': '', 'Email': '', 'Phone': '', 'Town': '', 'Position': '', 'Title': ''}
            for value in info_dict:
                if position_dict[value] == 0:
                    continue
                info_dict[value] = sheet.cell(row=i, column=position_dict[value]).value
                if info_dict[value] is not None:
                    if s in 'All SBs, TAs, Mayors, CityTown ' and value == 'Position':
                        info_dict['Title'] = info_dict['Position']
                        if 'Town Administrator' in info_dict['Position'] or 'Councilor' in info_dict['Position']:
                            if 'Councilor' in info_dict['Position']:
                                info_dict['Position'] = 'City & Town Councils'
                        else:
                            info_dict[value] = 'Muni officials: %s' % info_dict['Town']
                    if s in 'City and Town Councils' and value == 'Position':
                        info_dict['Title'] = info_dict[value]
                        info_dict[value] = 'City and Town Councils'
                    if 'Clerks' in s and value == 'Position':
                        info_dict[value] = 'Clerks'
                    if 'Public School Principals' in s and value == 'Position':
                        info_dict['Title'] = info_dict[value]
                        info_dict[value] = 'Public School Principals'
                    if 'Town Administrators' in s and value == 'Position':
                        info_dict['Title'] = info_dict[value] + ' ' + info_dict['Town']
                        info_dict[value] = 'Town Administrators'
                    if 'Water District Officials' in s and value == 'Position':
                        info_dict['Title'] = info_dict['Position'] + ' ' + info_dict['Town']
                        info_dict[value] = 'Water District Officials'
                    if 'Boards of Health' in s and value == 'Position':
                        info_dict['Title'] = 'Town'
                        info_dict[value] = 'Boards of Health'
                    if 'School Committees' in s and value == 'Position':
                        info_dict['Title'] = info_dict[value] + ', ' + info_dict['Town']
                        info_dict[value] = 'School Committees'
                    if 'Superintendents' in s and value == 'Position':
                        info_dict['Title'] = info_dict[value] + ', ' + info_dict['Town']
                        info_dict[value] = 'Superintendents'
                    if 'Fire Chief' in info_dict[value] and value == 'Position':
                        info_dict['Title'] = info_dict[value] + ', ' + info_dict['Town']
                        info_dict[value] = 'Fire Chiefs'
                    if 'Librarians' in s and value == 'Position':
                        info_dict['Title'] = info_dict['Town']
                        info_dict[value] = 'Public Libraries'
                    if 'Chief of Police' in info_dict[value] and value == 'Position':
                        info_dict['Title'] = info_dict[value] + ', ' + info_dict['Town']
                        info_dict[value] = 'Police Chiefs'
            positions_list = ['Muni officials:', 'Police Chiefs', 'Superintendents', 'Fire Chiefs', 'Public Libraries', 'Public School Principals', 'Boards of Health', 'City & Town Councils', 'Clerks', 'Town Administrators', 'School Committees', 'Water District Officials']
            position_boolean = False
            for p in positions_list:
                if str(info_dict['Position']) not in p:
                    position_boolean = True
            if position_boolean is False:
                info_dict['Position'] = 'Muni officials: %s' % info_dict['Town']
            dict_list.append(info_dict)

    return dict_list


def add_email_phone(driver, upload_list, driver2, skipped_list):
    name_elem_num = ''
    for d in upload_list:
        try:
            finish = True
            if d['Email'] != '' and d['Email'] != ' ':
                name = d['Name']
                name_field = driver.find_element_by_id('call-subject')
                #name_field.clear()
                for i in list(name):
                    name_field.send_keys(i)
                time.sleep(2)
                try:
                    element = driver.find_element_by_xpath('//*[@id="call-list"]/table/tbody/tr[1]/td[1]/span')
                    driver.execute_script("$(arguments[0]).click();", element)
                except Exception as exc:
                    print('Error: %s. %s will be skipped.' % (exc, d['Name']))
                    skipped_list.write('\n%s has been skipped.' % d['Name'])
                    driver.refresh()
                    driver.find_element_by_id('call-log-tab').click()
                    continue
                try:
                    name_elem_num = get_constituent_number(driver)
                except IndexError:
                    finish = False
                if finish is True:
                    positions_list = ['Muni officials:', 'Police Chiefs', 'Superintendents', 'Fire Chiefs',
                                      'Public Libraries', 'Public School Principals', 'Boards of Health',
                                      'City & Town Councils', 'Clerks', 'Town Administrators', 'School Committees',
                                      'Water District Officials']
                    if 'Muni officials' in d['Position'] or d['Position'] in positions_list:
                        add_constituent_groups(d, driver2, name_elem_num)
                    driver2.get('https://communityfluency.com/office/constituents/' + name_elem_num)
                    driver2.find_element_by_xpath('//*[@id="main"]/div[1]/a/button').click()
                    soup = bs4.BeautifulSoup(driver2.page_source, 'html.parser')
                    primary_email = soup.select('#contact_form > table > tbody > tr:nth-child(6) > td:nth-child(2) > div:nth-child(1) > input')[0]
                    primary_phone = soup.select('#contact_form > table > tbody > tr:nth-child(7) > td:nth-child(2) > div:nth-child(1) > input')[0]
                    if d['Email'] is not None and d['Email'] != '':
                        if primary_email.get('value').lower() != d['Email'].lower():
                            if primary_email.get('value') == '':
                                email_input = driver2.find_element_by_xpath('//*[@id="contact_form"]/table/tbody/tr[6]/td[2]/div[1]/input')
                            else:
                                email_input = driver2.find_element_by_xpath('//*[@id="contact_form"]/table/tbody/tr[6]/td[2]/div[3]/input[1]')
                            email_input.send_keys(d['Email'])
                    if d['Phone'] != '' and d['Phone'] is not None:
                        if primary_phone.get('value') == '' or primary_phone.get('value') == d['Phone']:
                            phone_input_elem = driver2.find_element_by_xpath('//*[@id="contact_form"]/table/tbody/tr[7]/td[2]/div[3]/input[1]')
                        else:
                            phone_input_elem = driver2.find_element_by_xpath('//*[@id="contact_form"]/table/tbody/tr[7]/td[2]/div[3]/input[1]')
                        if 'Ext' not in d['Phone']:
                            for i in list(d['Phone']):
                                if i.isdigit() is True or i == '-' or i == '.' or i == '(' or i == ')':
                                    phone_input_elem.send_keys(i)
                        else:
                            phone_input_elem.send_keys(d['Phone'])
                    driver2.find_element_by_xpath('//*[@id="contact_form"]/div[2]/input[2]').click()
                    print('%s file completed' % d['Name'])
                driver.refresh()
                driver.find_element_by_id('call-log-tab').click()
        except Exception as exc:
            print('Error: %s. %s will be skipped.' % (exc, d['Name']))
            skipped_list.write('\n%s has been skipped.' % d['Name'])
            driver.refresh()
            driver.find_element_by_id('call-log-tab').click()
            continue


def get_constituent_number(driver):
    html_text = driver.page_source
    soup = bs4.BeautifulSoup(html_text, 'html.parser')
    elems = soup.select('.ml-2')
    name_elem = elems[5]
    name_elem_num = ''
    start_value = ''
    append = False
    for i in list(str(name_elem)):
        start_value += i
        if start_value == '<label class="ml-2" for="add-person-':
            append = True
            continue
        if append is True:
            if i == '>':
                break
            name_elem_num += i
    return name_elem_num[:-1]


def add_constituent_groups(constituent_dict, driver, name_elem_num):
    driver.get('https://communityfluency.com/office/constituents/%s/category/94/new' % name_elem_num)
    position = constituent_dict['Position']
    soup = bs4.BeautifulSoup(driver.page_source, 'html.parser')
    id_dict = {'Boards of Health': '3711', 'Citizens Legislative Seminar': '4133', 'City & Town Councils': '3707', 'Clerks': '3706', 'Farmers': '4844', 'Fire Chiefs': '5116', 'Health Departments': '5113', 'Muni officials: Amherst': '3704', 'Muni officials: Bernardston': '3680', 'Muni officials: Colrain': '3681', 'Muni officials: Deerfield': '3682', 'Muni officials: Erving': '3683', 'Muni officials: Gill': '3684', 'Muni officials: Greenfield': '3685', 'Muni officials: Hadley': '3697', 'Muni officials: Hatfield': '3698', 'Muni officials: Leverett': '3686', 'Muni officials: Leyden': '3687', 'Muni officials: Montague': '3688', 'Muni officials: New Salem': '3689', 'Muni officials: Northampton': '3699', 'Muni officials: Northfield': '3690', 'Muni officials: Orange': '3691', 'Muni officials: Pelham': '3700', 'Muni officials: Royalston': '3702', 'Muni officials: Shutesbury': '3692', 'Muni officials: South Hadley': '3701', 'Muni officials: Sunderland': '3693', 'Muni officials: Warwick': '3694', 'Muni officials: Wendell': '3695', 'Muni officials: Whately': '3696', 'Police Chiefs': '5115', 'Public Libraries': '5118', 'Public School Principals': '3708', 'Public Works': '5119', 'School Committees': '3712', 'Selectboards': '3705', 'Superintendents': '5114', 'Test: Personal': '5166', 'Test: Work': '5165', "The People's Office": '3679', 'Town Administrators': '3709', 'Under 18': '4748', 'Veterans Service Officers': '5117', 'Water District Officials': '3710'}
    id_num = id_dict[position]
    issue_elem = soup.find(id=id_num)
    issue_elem_name = issue_elem.get('name')
    if 'keep' not in issue_elem_name:
        driver.find_element_by_id(id_num).click()
        if constituent_dict['Title'] != '':
            driver.find_element_by_name('title_' + id_num).send_keys(constituent_dict['Title'])
    driver.find_element_by_xpath('//*[@id="main"]/form/div[2]/input[2]').click()


main()
