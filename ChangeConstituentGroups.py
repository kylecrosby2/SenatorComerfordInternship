#! python3
# ChangeConstituentGroups.py - Edit the constituent groups of officials in CF from the spreadsheet.

from selenium import webdriver
import bs4
import time
import openpyxl as excel


def main():
    path = "C:/Users/kerry/Documents/chromedriver.exe"
    driver = webdriver.Chrome(path)
    driver2 = webdriver.Chrome(path)
    login_cf(driver)
    login_cf(driver2)
    skipped_file = open('SkippedSBList1.txt', 'a', encoding='utf-8')
    wb = excel.load_workbook('EmailList11-24.xlsx')
    sheet = wb['All SBs, TAs, Mayors, CityTown ']
    member_dict_list = get_data_from_sheet(sheet)
    cf_page(member_dict_list, driver, driver2, skipped_file)
    skipped_file.close()


# Logs into cf with the webdriver.
def login_cf(driver):
    driver.get("https://communityfluency.com/office")
    email = driver.find_element_by_id('email')
    password = driver.find_element_by_id('password')
    # Must input your email and password into the quotations below.
    email.send_keys('akylecrosby@gmail.com')
    password.send_keys('$omewhere391Flag')
    sign_in = driver.find_elements_by_xpath('//button[@type="submit"]')[0]
    sign_in.click()
    driver.find_element_by_id('call-log-tab').click()


def get_data_from_sheet(sheet):
    member_dict_list = []
    for i in range(2, 147):
        member_dict = {'Name': sheet.cell(row=i, column=2).value, 'Email': sheet.cell(row=i, column=4).value, 'Position': sheet.cell(row=i, column=3).value, 'Town': sheet.cell(row=i, column=1).value}
        if 'Selectboard' in str(member_dict['Position']):
            member_dict_list.append(member_dict)
    return member_dict_list


def cf_page(member_dict_list, driver, driver2, skipped_file):
    for d in member_dict_list:
        try:
            driver.refresh()
            driver.find_element_by_id('call-log-tab').click()
            name_elem_num = get_constituent_number(driver, d, skipped_file)
            add_to_constituent_groups(d, driver2, name_elem_num)
        except Exception as exc:
            print('Error: %s. %s will be skipped.' % (exc, d['Name']))
            skipped_file.write(str(d))
            skipped_file.write('\n\n')
            continue


def get_constituent_number(driver, member_dict, skipped_file):
    finish = True
    name_elem_num = ''
    name = member_dict['Name']
    name_field = driver.find_element_by_id('call-subject')
    # name_field.clear()
    for i in list(name):
        name_field.send_keys(i)
    time.sleep(2)
    try:
        element = driver.find_element_by_xpath('//*[@id="call-list"]/table/tbody/tr[1]/td[1]/span')
        driver.execute_script("$(arguments[0]).click();", element)
    except Exception as exc:
        print('Error: %s. %s will be skipped.' % (exc, name))
        skipped_file.write(str(member_dict))
        driver.refresh()
        driver.find_element_by_id('call-log-tab').click()
        finish = False
    if finish is True:
        html_text = driver.page_source
        soup = bs4.BeautifulSoup(html_text, 'html.parser')
        elems = soup.select('.ml-2')
        name_elem = elems[5]
        name_elem_num = ''
        start_value = ''
        append = False
        for i in list(str(name_elem)):
            start_value += i
            if start_value == '<label class="ml-2" for="add-person-':
                append = True
                continue
            if append is True:
                if i == '>':
                    break
                name_elem_num += i
    return name_elem_num[:-1]


def add_to_constituent_groups(member_dict, driver, name_elem_num):
    driver.get('https://communityfluency.com/office/constituents/%s/category/94/new' % name_elem_num)
    soup = bs4.BeautifulSoup(driver.page_source, 'html.parser')
    #con_group = 'Muni officials: ' + member_dict['Town'].strip()
    con_group = 'Selectboards'
    con_group_num_dict = {'Boards of Health': '3711', 'Citizens Legislative Seminar': '4133', 'City & Town Councils': '3707', 'Clerks': '3706', 'Economic Development Contacts': '5318', 'Farmers': '4844', 'Fire Chiefs': '5116', 'Health Departments': '5113', 'Muni officials: Amherst': '3704', 'Muni officials: Bernardston': '3680', 'Muni officials: Colrain': '3681', 'Muni officials: Deerfield': '3682', 'Muni officials: Erving': '3683', 'Muni officials: Gill': '3684', 'Muni officials: Greenfield': '3685', 'Muni officials: Hadley': '3697', 'Muni officials: Hatfield': '3698', 'Muni officials: Leverett': '3686', 'Muni officials: Leyden': '3687', 'Muni officials: Montague': '3688', 'Muni officials: New Salem': '3689', 'Muni officials: Northampton': '3699', 'Muni officials: Northfield': '3690', 'Muni officials: Orange': '3691', 'Muni officials: Pelham': '3700', 'Muni officials: Royalston': '3702', 'Muni officials: Shutesbury': '3692', 'Muni officials: South Hadley': '3701', 'Muni officials: Sunderland': '3693', 'Muni officials: Warwick': '3694', 'Muni officials: Wendell': '3695', 'Muni officials: Whately': '3696', 'Police Chiefs': '5115', 'Public Libraries': '5118', 'Public School Principals': '3708', 'Public Works': '5119', 'School Committees': '3712', 'Selectboards': '3705', 'Superintendents': '5114', 'Test: Personal': '5166', 'Test: Work': '5165', "The People's Office": '3679', 'Town Administrators': '3709', 'Under 18': '4748', 'Veterans Service Officers': '5117', 'Water District Officials': '3710'}
    checkbox_elem = soup.find(id=con_group_num_dict[con_group])
    if 'keep' not in checkbox_elem.get('name'):
        driver.find_element_by_xpath('//*[@id="%s"]' % con_group_num_dict[con_group]).click()
        driver.find_element_by_xpath('//*[@id="pivot_%s"]/div[1]/input' % con_group_num_dict[con_group]).send_keys(
            '%s, %s' % (member_dict['Position'], member_dict['Town']))
    time.sleep(1)
    driver.find_element_by_xpath('//*[@id="main"]/form/div[2]/input[2]').click()
    time.sleep(1)


main()
